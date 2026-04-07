"""
Export service for generating CSV, Excel, and JSON exports.

Supports multi-sheet Excel export with proper formatting for:
- Summary statistics
- Forms data
- Emails data
- Invoices data
"""

import csv
import io
import json
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.logging import audit_logger, get_logger
from app.db.models import ExtractionRecordDB
from app.db.repositories import RecordRepository
from app.models.schemas import ExportRequest

logger = get_logger(__name__)

# Column definitions per record type
FORM_COLUMNS = [
    "Client_Name", "Email", "Phone", "Company",
    "Service_Interest", "Priority", "Message", "Date", "Status", "Confidence"
]
EMAIL_COLUMNS = [
    "Client_Name", "Email", "Company", "Service_Interest",
    "Invoice_Number", "Amount", "Message", "Date", "Status", "Confidence"
]
INVOICE_COLUMNS = [
    "Invoice_Number", "Client_Name", "Amount", "VAT",
    "Total_Amount", "Date", "Status", "Confidence"
]


class ExportService:
    """
    Service for exporting records to various formats.

    Supports:
    - CSV export
    - Excel (XLSX) export
    - JSON export
    """

    def __init__(self, repository: RecordRepository):
        """
        Initialize service with repository.

        Args:
            repository: RecordRepository for data access.
        """
        self.repository = repository

    async def export_records(
        self, request: ExportRequest
    ) -> tuple[bytes, str, str, list[str]]:
        """
        Export records to the requested format.

        Args:
            request: ExportRequest with format and filters.

        Returns:
            Tuple of (file_bytes, filename, content_type, exported_record_ids).
            exported_record_ids contains IDs of records marked as 'exported'.

        Raises:
            ValueError: If no records to export or unsupported format.
        """
        # Get records to export
        records = await self.repository.get_exportable_records(
            record_ids=request.record_ids,
            include_rejected=request.include_rejected,
        )

        if not records:
            raise ValueError("No records to export")

        # Convert to flat dictionaries following the template schema
        data = [self._flatten_record(r) for r in records]

        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

        # Export based on format
        if request.format == "csv":
            content, filename = self._export_csv(data, timestamp)
            content_type = "text/csv; charset=utf-8"
        elif request.format == "xlsx":
            content, filename = self._export_xlsx(data, timestamp)
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        elif request.format == "json":
            content, filename = self._export_json(data, timestamp)
            content_type = "application/json; charset=utf-8"
        else:
            raise ValueError(f"Unsupported export format: {request.format}")

        # Log the export
        audit_logger.log_export(
            export_format=request.format,
            record_count=len(records),
            destination=filename,
        )

        logger.info(
            "records_exported",
            format=request.format,
            count=len(records),
            filename=filename,
        )

        # Mark records as exported (only approved, not edited or rejected)
        # Edited records are included in exports but keep their status
        # since they haven't been formally approved yet
        # Collect IDs of records that were marked as exported for auto-sync
        exported_ids: list[str] = []
        for record in records:
            if record.status == "approved":
                record.status = "exported"
                record.updated_at = datetime.now(UTC)
                await self.repository.update(record)
                exported_ids.append(str(record.id))

        # Commit changes so background sync can see them
        await self.repository.commit()

        return content, filename, content_type, exported_ids

    def _flatten_record(self, record: ExtractionRecordDB) -> dict[str, Any]:
        """
        Flatten a record for export following the template schema.

        The template from data_extraction_template.csv has columns:
        Type, Source, Date, Client_Name, Email, Phone, Company,
        Service_Interest, Amount, VAT, Total_Amount, Invoice_Number,
        Priority, Message

        Args:
            record: ExtractionRecordDB to flatten.

        Returns:
            Flat dictionary matching template schema.
        """
        # Use edited_data if available, otherwise extracted_data
        data = record.final_data

        # Base fields
        flat: dict[str, Any] = {
            "Type": record.record_type,
            "Source": record.source_file,
            "Date": record.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "Status": record.status,
            "Confidence": f"{record.confidence_score:.2%}",
        }

        # Type-specific fields following data_extraction_template.csv
        if record.record_type == "FORM":
            flat.update(
                {
                    "Client_Name": data.get("full_name"),
                    "Email": data.get("email"),
                    "Phone": data.get("phone"),
                    "Company": data.get("company"),
                    "Service_Interest": data.get("service_interest"),
                    "Message": data.get("message"),
                    "Priority": data.get("priority"),
                    "Amount": None,
                    "VAT": None,
                    "Total_Amount": None,
                    "Invoice_Number": None,
                }
            )
        elif record.record_type == "EMAIL":
            flat.update(
                {
                    "Client_Name": data.get("sender_name"),
                    "Email": data.get("sender_email"),
                    "Phone": data.get("phone"),
                    "Company": data.get("company"),
                    "Service_Interest": data.get("service_interest"),
                    "Message": self._truncate_text(data.get("body"), 500),
                    "Priority": None,
                    "Amount": data.get("invoice_amount"),
                    "VAT": None,
                    "Total_Amount": None,
                    "Invoice_Number": data.get("invoice_number"),
                }
            )
        elif record.record_type == "INVOICE":
            flat.update(
                {
                    "Client_Name": data.get("client_name"),
                    "Email": None,
                    "Phone": None,
                    "Company": None,
                    "Service_Interest": None,
                    "Message": data.get("notes"),
                    "Priority": None,
                    "Amount": data.get("net_amount"),
                    "VAT": data.get("vat_amount"),
                    "Total_Amount": data.get("total_amount"),
                    "Invoice_Number": data.get("invoice_number"),
                }
            )

        return flat

    def _truncate_text(self, text: str | None, max_length: int) -> str | None:
        """Truncate text to max length with ellipsis."""
        if text is None:
            return None
        if len(text) <= max_length:
            return text
        return text[: max_length - 3] + "..."

    def _export_csv(
        self, data: list[dict[str, Any]], timestamp: str
    ) -> tuple[bytes, str]:
        """
        Export data to CSV format.

        Args:
            data: List of flattened record dictionaries.
            timestamp: Timestamp string for filename.

        Returns:
            Tuple of (CSV bytes, filename).
        """
        output = io.StringIO()

        if data:
            # Define column order following template
            fieldnames = [
                "Type",
                "Source",
                "Date",
                "Status",
                "Confidence",
                "Client_Name",
                "Email",
                "Phone",
                "Company",
                "Service_Interest",
                "Amount",
                "VAT",
                "Total_Amount",
                "Invoice_Number",
                "Priority",
                "Message",
            ]

            writer = csv.DictWriter(
                output, fieldnames=fieldnames, extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(data)

        # Add UTF-8 BOM for Excel compatibility with Greek characters
        csv_bytes = b'\xef\xbb\xbf' + output.getvalue().encode("utf-8")
        return csv_bytes, f"ellincrm_export_{timestamp}.csv"

    def _export_xlsx(
        self, data: list[dict[str, Any]], timestamp: str
    ) -> tuple[bytes, str]:
        """
        Export data to Excel format with multi-sheet organization.

        Creates 4 sheets:
        - Summary: Statistics and totals
        - Forms: Contact form submissions
        - Emails: Email extractions
        - Invoices: Invoice data with financial totals

        Args:
            data: List of flattened record dictionaries.
            timestamp: Timestamp string for filename.

        Returns:
            Tuple of (Excel bytes, filename).
        """
        # Separate data by type
        forms_data = [r for r in data if r["Type"] == "FORM"]
        emails_data = [r for r in data if r["Type"] == "EMAIL"]
        invoices_data = [r for r in data if r["Type"] == "INVOICE"]

        # Create workbook
        wb = Workbook()

        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        title_font = Font(bold=True, size=14, color="2F5496")
        subtitle_font = Font(bold=True, size=11)

        status_fills = {
            "pending": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "approved": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "rejected": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "exported": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
            "edited": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        }

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 1. Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(
            ws_summary, data, forms_data, emails_data, invoices_data,
            title_font, subtitle_font, header_font, header_fill, thin_border
        )

        # 2. Forms Sheet
        ws_forms = wb.create_sheet("Forms")
        self._create_data_sheet(
            ws_forms, forms_data, FORM_COLUMNS, "Contact Forms",
            header_font, header_fill, header_alignment, status_fills, thin_border
        )

        # 3. Emails Sheet
        ws_emails = wb.create_sheet("Emails")
        self._create_data_sheet(
            ws_emails, emails_data, EMAIL_COLUMNS, "Email Extractions",
            header_font, header_fill, header_alignment, status_fills, thin_border
        )

        # 4. Invoices Sheet
        ws_invoices = wb.create_sheet("Invoices")
        self._create_data_sheet(
            ws_invoices, invoices_data, INVOICE_COLUMNS, "Invoices",
            header_font, header_fill, header_alignment, status_fills, thin_border,
            currency_columns=["Amount", "VAT", "Total_Amount"]
        )

        output = io.BytesIO()
        wb.save(output)

        return output.getvalue(), f"ellincrm_export_{timestamp}.xlsx"

    def _create_summary_sheet(
        self,
        ws,
        data: list[dict],
        forms_data: list[dict],
        emails_data: list[dict],
        invoices_data: list[dict],
        title_font: Font,
        subtitle_font: Font,
        header_font: Font,
        header_fill: PatternFill,
        border: Border,
    ) -> None:
        """Create the summary sheet with statistics."""
        # Title
        ws["A1"] = "EllinCRM Data Export Summary"
        ws["A1"].font = title_font
        ws.merge_cells("A1:D1")

        ws["A2"] = f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws["A2"].font = Font(italic=True, color="666666")

        # Record counts section
        ws["A4"] = "Record Counts"
        ws["A4"].font = subtitle_font

        summary_data = [
            ("Total Records", len(data)),
            ("Forms", len(forms_data)),
            ("Emails", len(emails_data)),
            ("Invoices", len(invoices_data)),
        ]

        for i, (label, value) in enumerate(summary_data, start=5):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].border = border
            ws[f"B{i}"].border = border

        # Status breakdown
        ws["A10"] = "Status Breakdown"
        ws["A10"].font = subtitle_font

        status_counts = {}
        for record in data:
            status = record.get("Status", "unknown")
            status_counts[status] = status_counts.get(status, 0) + 1

        row = 11
        for status, count in sorted(status_counts.items()):
            ws[f"A{row}"] = status.capitalize()
            ws[f"B{row}"] = count
            ws[f"A{row}"].border = border
            ws[f"B{row}"].border = border
            row += 1

        # Financial summary (invoices only)
        if invoices_data:
            ws[f"A{row + 1}"] = "Financial Summary (Invoices)"
            ws[f"A{row + 1}"].font = subtitle_font

            total_amount = sum(
                float(r.get("Amount") or 0) for r in invoices_data
            )
            total_vat = sum(
                float(r.get("VAT") or 0) for r in invoices_data
            )
            total_with_vat = sum(
                float(r.get("Total_Amount") or 0) for r in invoices_data
            )

            financial_data = [
                ("Net Amount", f"€{total_amount:,.2f}"),
                ("VAT (24%)", f"€{total_vat:,.2f}"),
                ("Total Amount", f"€{total_with_vat:,.2f}"),
            ]

            for i, (label, value) in enumerate(financial_data, start=row + 2):
                ws[f"A{i}"] = label
                ws[f"B{i}"] = value
                ws[f"A{i}"].border = border
                ws[f"B{i}"].border = border
                if label == "Total Amount":
                    ws[f"A{i}"].font = Font(bold=True)
                    ws[f"B{i}"].font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_data_sheet(
        self,
        ws,
        data: list[dict],
        columns: list[str],
        title: str,
        header_font: Font,
        header_fill: PatternFill,
        header_alignment: Alignment,
        status_fills: dict[str, PatternFill],
        border: Border,
        currency_columns: list[str] | None = None,
    ) -> None:
        """Create a data sheet with headers and formatted data."""
        if not data:
            ws["A1"] = f"No {title.lower()} data to display"
            return

        currency_columns = currency_columns or []

        # Write headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " "))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        for row_idx, record in enumerate(data, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = record.get(col_name)

                # Format currency values
                if col_name in currency_columns and value is not None:
                    try:
                        value = f"€{float(value):,.2f}"
                    except (ValueError, TypeError):
                        pass

                # Truncate long messages
                if col_name == "Message" and value and len(str(value)) > 100:
                    value = str(value)[:97] + "..."

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(col_name == "Message"))

                # Apply status color
                if col_name == "Status" and value:
                    status_key = str(value).lower()
                    if status_key in status_fills:
                        cell.fill = status_fills[status_key]

        # Auto-fit column widths (approximate)
        column_widths = {
            "Client_Name": 20, "Email": 25, "Phone": 15, "Company": 20,
            "Service_Interest": 18, "Priority": 10, "Message": 40,
            "Date": 18, "Status": 12, "Confidence": 12,
            "Invoice_Number": 15, "Amount": 12, "VAT": 12, "Total_Amount": 14,
        }
        for col_idx, col_name in enumerate(columns, start=1):
            width = column_widths.get(col_name, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    def _export_json(
        self, data: list[dict[str, Any]], timestamp: str
    ) -> tuple[bytes, str]:
        """
        Export data to JSON format.

        Args:
            data: List of flattened record dictionaries.
            timestamp: Timestamp string for filename.

        Returns:
            Tuple of (JSON bytes, filename).
        """
        json_output = {
            "exported_at": datetime.now(UTC).isoformat(),
            "record_count": len(data),
            "records": data,
        }

        return (
            json.dumps(json_output, ensure_ascii=False, indent=2).encode("utf-8"),
            f"ellincrm_export_{timestamp}.json",
        )
